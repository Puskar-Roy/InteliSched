import customtkinter as ctk
import tkinter as tk
from time import strftime
import subprocess
import pyrebase
from tkinter import messagebox
from tkinter import *
from tkinter import filedialog
import openpyxl
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore

cred = credentials.Certificate("firebase_clouddb.json")
app = firebase_admin.initialize_app(cred,name="admin_panel")
db = firestore.client(app=app)


firebaseConfig = {
  'apiKey': "AIzaSyDlwmXi6-Y8fPhURBAdJybXlJuSbvew-PE",
  'authDomain': "formintellisched.firebaseapp.com",
  'databaseURL': "https://formintellisched-default-rtdb.firebaseio.com",
  'projectId': "formintellisched",
  'storageBucket': "formintellisched.appspot.com",
  'messagingSenderId': "725143468756",
  'appId': "1:725143468756:web:920f84d70991093482d941",
  'measurementId': "G-SNSBZJ54Q2"
}
firebase = pyrebase.initialize_app(firebaseConfig)
auth = firebase.auth()
# db = firebase.database() its for real time data base



def handel_admin_details(db_id):
    # result = db.collection("Admin").document(db_id).get()
    # user_data = 0
    # if result.exists:
    #     user_data = result.to_dict() 

    result = db.collection("Admin").document(db_id).get()
    user_data = 0
    if result.exists:
        user_data = result.to_dict()           

    print(user_data["Name"])    
        
    def delete_page():
            for frame in main_frame.winfo_children():
                frame.destroy()

# ==========================================================================================Notice Page 
    def notice_page(event):
            delete_page()
            
            frame1 = ctk.CTkFrame(main_frame,width=1100,height=750)



            img1 = PhotoImage(file="logo1 (1).png")
            label = ctk.CTkLabel(frame1, image=img1,text="",height=10,width=10)
            label.place(x=35,y=16)



            def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


            mark = ctk.CTkLabel(frame1, 
                        font = ('calibri', 20, 'bold'))
            mark.place(x=910,y=16)
            time()


            lablex= ctk.CTkLabel(frame1,text="Welcome",
                                font=ctk.CTkFont(weight="bold",size=40,family="Verdana"))
            lablex.place(x=100,y=15)

            labley= ctk.CTkLabel(frame1,text=user_data["Name"],
                                font=ctk.CTkFont(weight="bold",size=27,family="Helvetica"))
            labley.place(x=30,y=70)

            lable= ctk.CTkLabel(frame1,text="Notice Board",
                                font=ctk.CTkFont(weight="bold",size=35))
            lable.place(x=445,y=160)
            frame1.pack(pady=20)



# =====================================================================================================Log Out


    def log_out():
            response = messagebox.askyesno(title="Log Out",message="Confirm Exit?")
            if response == True:
                root.destroy()
            else:
                None

            subprocess.run(["python3","d_login_page.py"])    

                
            



# ==================================================================================================Admin Details Page 

    def home_page(event):
            delete_page()
            frame1 = ctk.CTkFrame(main_frame,width=1100,height=770)


            img1 = PhotoImage(file="logo1 (1).png")
            label = ctk.CTkLabel(frame1, image=img1,text="",height=10,width=10)
            label.place(x=35,y=16)



            def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


            mark = ctk.CTkLabel(frame1, 
                        font = ('calibri', 20, 'bold'))
            mark.place(x=910,y=16)
            time()


            lablex= ctk.CTkLabel(frame1,text="Welcome",
                                font=ctk.CTkFont(weight="bold",size=40,family="Verdana"))
            lablex.place(x=100,y=15)

            labley= ctk.CTkLabel(frame1,text=user_data["Name"],
                                font=ctk.CTkFont(weight="bold",size=27,family="Helvetica"))
            labley.place(x=30,y=70)


            lable= ctk.CTkLabel(frame1,text="Admin Details",
                                font=ctk.CTkFont(weight="bold",size=35,family="Verdana"))
            lable.place(x=405,y=140)


            lable1 = ctk.CTkLabel(frame1,text="Name : "+user_data["Name"],
                                font=ctk.CTkFont(weight="bold",size=25,family="Helvetica"))
            lable1.place(x=40,y=240)
            lable2 = ctk.CTkLabel(frame1,text="Email : "+user_data["Email"],
                                font=ctk.CTkFont(weight="bold",size=25,family="Helvetica"))
            lable2.place(x=550,y=240)

            # lable3 = ctk.CTkLabel(frame1,text="Subject Code : ",
            #                     font=ctk.CTkFont(weight="bold",size=25,family="Helvetica"))
            # lable3.place(x=40,y=315)

            # lable4 = ctk.CTkLabel(frame1,text="Depertment : ",
            #                     font=ctk.CTkFont(weight="bold",size=25,family="Helvetica"))
            # lable4.place(x=550,y=315)

            # lable6 = ctk.CTkLabel(frame1,text="Address : ",
            #                     font=ctk.CTkFont(weight="bold",size=25,family="Helvetica"))
            # lable6.place(x=40,y=390)

            frame1.pack(pady=20)



    # ---------------------------------------------------------------------------------Change Password Page




    def change_password():
        def update():
            if(cur_mail.get() == user_data["Password"]):
                if(new_pass.get()!= cur_mail.get()):
                    user_id = auth.sign_in_with_email_and_password(user_data["Email"],user_data["Password"])
                    idt = user_id["idToken"]
                    auth.delete_user_account(idt)
                    auth.create_user_with_email_and_password(user_data["Email"],new_pass.get())
                    db.collection("Admin").document(user_data["Admin Id"]).update({"Password":new_pass.get()})
                    messagebox.showinfo(title="Updtae",message="Update Email Successfull")
                    root.destroy()
                else:
                    messagebox.showerror(title="Updtae",message="Your new password must be different from current password")
                                  
            else:
                messagebox.showerror(title="Updtae",message="Your current password did not match,Try Again")

        root = ctk.CTk()
        root.geometry("600x400")

        root1 =ctk.CTkFrame(root,height=300,width=370)
        root1.pack(pady=45)


        label1 = ctk.CTkLabel(root1,text="Update Password",font=ctk.CTkFont(size=30,weight="bold"))
        label1.place(x=55,y=30)

        label1 = ctk.CTkLabel(root1,text="Enter Current Password",font=ctk.CTkFont(size=15,weight="bold"))
        label1.place(x=20,y=90)

        cur_mail = ctk.CTkEntry(root1,placeholder_text="Enter Current Password",
                                height=32,
                                width=300,
                                corner_radius=10,
                                border_color="#3468c9",
                                font=ctk.CTkFont(size=15,weight="bold"))
        cur_mail.place(x=20,y=120)


        label1 = ctk.CTkLabel(root1,text="Enter New Password",font=ctk.CTkFont(size=15,weight="bold"))
        label1.place(x=20,y=170)

        new_pass = ctk.CTkEntry(root1,placeholder_text="Enter New Password",
                                height=32,
                                width=300,
                                corner_radius=10,
                                border_color="#3468c9",
                                font=ctk.CTkFont(size=15,weight="bold"))
        new_pass.place(x=20,y=200)

        btnx = ctk.CTkButton(root1,text="Change",
                            font=ctk.CTkFont(size=15,weight="bold"),
                            command=update)
        btnx.place(y=250,x=180)
        root.mainloop()





    # ---------------------------------------------------------------------------------Change Email Page
    



    def change_email():
        def update():
            if(cur_mail.get() == user_data["Email"]):
                if(new_mail.get()!= cur_mail.get()):
                    user_id = auth.sign_in_with_email_and_password(user_data["Email"],user_data["Password"])
                    idt = user_id["idToken"]
                    auth.delete_user_account(idt)
                    auth.create_user_with_email_and_password(new_mail.get(),user_data["Password"])
                    db.collection("Admin").document(user_data["Admin Id"]).update({"Email":new_mail.get()})
                    messagebox.showinfo(title="Updtae",message="Update Email Successfull")
                    root.destroy()
                else:
                    messagebox.showerror(title="Updtae",message="Your new email must be different from current email")
                                  
            else:
                messagebox.showerror(title="Updtae",message="Your current email did not match,Try Again")




        root = ctk.CTk()
        root.geometry("600x400")
        root.title("Update Email")

        root1 =ctk.CTkFrame(root,height=300,width=370)
        root1.pack(pady=45)


        label1 = ctk.CTkLabel(root1,text="Update Email",font=ctk.CTkFont(size=30,weight="bold"))
        label1.place(x=95,y=30)

        label1 = ctk.CTkLabel(root1,text="Enter Current Email",font=ctk.CTkFont(size=15,weight="bold"))
        label1.place(x=20,y=90)

        cur_mail = ctk.CTkEntry(root1,placeholder_text="Enter Current Email",
                                height=32,
                                width=300,
                                corner_radius=10,
                                border_color="#3468c9",
                                font=ctk.CTkFont(size=15,weight="bold"))
        cur_mail.place(x=20,y=120)


        label1 = ctk.CTkLabel(root1,text="Enter New Email",font=ctk.CTkFont(size=15,weight="bold"))
        label1.place(x=20,y=170)

        new_mail = ctk.CTkEntry(root1,placeholder_text="Enter New Email",
                                height=32,
                                width=300,
                                corner_radius=10,
                                border_color="#3468c9",
                                font=ctk.CTkFont(size=15,weight="bold"))
        new_mail.place(x=20,y=200)

        btnx = ctk.CTkButton(root1,text="Change",
                            font=ctk.CTkFont(size=15,weight="bold"),
                            command=update)
        btnx.place(y=250,x=180)
        root.mainloop()

# ===================================================================================update profile page


    def edite_page(event):
            delete_page()
            frame1 = ctk.CTkFrame(main_frame,width=1100,height=750)
            img1 = PhotoImage(file="logo1 (1).png")
            label = ctk.CTkLabel(frame1, image=img1,text="",height=10,width=10)
            label.place(x=35,y=16)
            def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


            mark = ctk.CTkLabel(frame1, 
                        font = ('calibri', 20, 'bold'))
            mark.place(x=910,y=16)
            time()


            lablex= ctk.CTkLabel(frame1,text="Welcome",
                                font=ctk.CTkFont(weight="bold",size=40,family="Verdana"))
            lablex.place(x=100,y=15)

            labley= ctk.CTkLabel(frame1,text="",
                                font=ctk.CTkFont(weight="bold",size=27))
            labley.place(x=30,y=70)

            l1 =ctk.CTkLabel(frame1,text="Update Profile",
                            font=ctk.CTkFont(weight="bold",size=35,family="Verdana"))
            l1.place(x=415,y=160)

            l2 =ctk.CTkLabel(frame1,text="Update :  ",
                            font=ctk.CTkFont(weight="bold",size=20))
            l2.place(x=420,y=250)

            bt1 =ctk.CTkButton(frame1,
                            text="Email",
                            font=ctk.CTkFont(weight="bold",size=20),
                            command=change_email)
            bt1.place(x=520,y=248)


            l3 =ctk.CTkLabel(frame1,text="Update :  ",
                            font=ctk.CTkFont(weight="bold",size=20))
            l3.place(x=420,y=310)

            bt2 =ctk.CTkButton(frame1,
                            text="Password",
                            font=ctk.CTkFont(weight="bold",size=20),
                            command=change_password)
            bt2.place(x=520,y=310)

            frame1.pack(pady=20)



    # ============================================================= Import Excel students
    
    def select_excel_file_stu():
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")]) 
            print(file_path)   
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook['Sheet1']
            names = []
            rollnumbers = []
            departments = []
            emails = []
            passwords = []
            count = 0
            for row in worksheet.iter_rows(values_only=True):
                names.append(row[0])
                rollnumbers.append(row[1])
                departments.append(row[2])
                emails.append(row[3])
                passwords.append(row[4])
                print(count," Iteration \n")
                print(names[count])
                print(rollnumbers[count])
                print(departments[count])
                print(emails[count])
                print(passwords[count])
                print("_"*50)
                
                user = auth.create_user_with_email_and_password(email=emails[count],password=passwords[count])
                data = {"Role":"Student","Name":names[count],"Depertment":departments[count],"Roll Number":rollnumbers[count],"Email":emails[count],"Password":passwords[count],"Phone Number":"nill"}
                # db.child("Students").child(rollnumbers[count]).set(data) for add data in real time firebase
                db.collection("Student").document(rollnumbers[count]).set(data)  #for add data in cloud store

                count+=1

            messagebox.showerror(title="Import Details",message="Import Details Successfully")

        except:
            messagebox.showerror(title="Import Details",message="Something went wrong please try again")
             

    # ============================================================= Import Excel faculty
    
    def select_excel_file_fac():
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")]) 
            print(file_path)   
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook['Sheet1']
            names = []
            rollnumbers = []
            departments = []
            emails = []
            passwords = []
            count = 0
            for row in worksheet.iter_rows(values_only=True):
                names.append(row[0])
                rollnumbers.append(row[1])
                departments.append(row[2])
                emails.append(row[3])
                passwords.append(row[4])
                print(count," Iteration \n")
                print(names[count])
                print(rollnumbers[count])
                print(departments[count])
                print(emails[count])
                print(passwords[count])
                print("_"*50)
                
                user = auth.create_user_with_email_and_password(email=emails[count],password=passwords[count])
                data = {"Role":"Faculty","Name":names[count],"Depertment":departments[count],"Faculty Id":rollnumbers[count],"Email":emails[count],"Password":passwords[count],"Phone Number":"nill"}
                # db.child("Students").child(rollnumbers[count]).set(data) for add data in real time firebase
                db.collection("Faculties").document(rollnumbers[count]).set(data)  #for add data in cloud store

                count+=1

            messagebox.showerror(title="Import Details",message="Import Details Successfully")

        except:
            messagebox.showerror(title="Import Details",message="Something went wrong please try again")
             
             
                 
    # -----------------------------------------------------------------------------------------------------   Registration Page

        
    def register_page_fac():

        def register():
            username = ph_Num.get()
            password = passw.get()
            dept = optionmenu_1.get()
            name = namee.get()
            phone_numb =phone.get()
            fac_id = code_Sub.get()
            if len(username) == 0 or len(password) == 0 or len(name) == 0 or len(dept) == 0 or len(fac_id) == 0 or len(phone_numb) == 0:
                messagebox.showerror(title="Sign Up",message="Plese Fill All Details")
            else:
                try:
                    user = auth.create_user_with_email_and_password(email=username,password=password)
                    data = {"Role":"Faculty","Name":name,"Depertment":dept,"Faculty Id":fac_id,"Email":username,"Password":password,"Phone Number":phone_numb}
                    # db.child("Faculties").child(fac_id).set(data)
                    db.collection("Faculties").document(fac_id).set(data)  #for add data in cloud store

                    messagebox.showinfo(title="Sign Up",message="Registration Successful, Now you can login with your email and password !")
                    root.destroy()
                except:
                    messagebox.showerror(title="Sign Up",message="Invalid email or password, Try Again !")
                    root.destroy()


        root = ctk.CTk()
        root.geometry("1100x700")
        root.title("Sign Up")
        root1 = ctk.CTkFrame(root,height=600,width=1000)
        root1.pack(pady=50)
        label1 = ctk.CTkLabel(root1,text="Register Facutly",font=ctk.CTkFont(size=40,weight="bold"))
        label1.place(x=330,y=45)

        def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


        mark = ctk.CTkLabel(root1, 
                        font = ('calibri', 20, 'bold'))
        mark.place(x=870,y=65)
        time()

        lable2 = ctk.CTkLabel(root1,text="Faculty Name",
                                        font=ctk.CTkFont(size=20,weight="bold"))
        lable2.place(x=110,y=155) 
        namee = ctk.CTkEntry(root1,placeholder_text="Enter Name", 
                                        width=360,
                                        height=40,
                                        font=ctk.CTkFont(size=15,weight="bold"),
                                        border_color="#3468c9",
                                        corner_radius=15)#check box
        namee.place(x=110,y=192)


        lable3 = ctk.CTkLabel(root1,text="Enter Email",
                                        font=ctk.CTkFont(size=20,weight="bold"))
        lable3.place(x=520,y=160) 

        ph_Num = ctk.CTkEntry(root1,placeholder_text="Enter Email", 
                                        width=360,
                                        height=40,
                                        font=ctk.CTkFont(size=15,weight="bold"),
                                        border_color="#3468c9",
                                        corner_radius=15)#check box
        ph_Num.place(x=520,y=192)



        lable4 = ctk.CTkLabel(root1,text="Enter Department",
                                        font=ctk.CTkFont(size=20,weight="bold"))
        lable4.place(x=115,y=255) 
        optionmenu_1 = ctk.CTkOptionMenu(root1,dynamic_resizing=False,
                                                        values=["CSE","IT","ECE","EE","ME","CE"],
                                                        width=360,
                                                        height=35,
                                                        dropdown_hover_color="#2f4978",
                                                        font=ctk.CTkFont(size=18,weight="bold"),
                                                        dropdown_font=ctk.CTkFont(size=20,weight="bold"))
        optionmenu_1.place(x=110,y=292)

        lable5 = ctk.CTkLabel(root1,text="Enter Subject-Code",
                                        font=ctk.CTkFont(size=20,weight="bold"))
        lable5.place(x=520,y=255) 
        code_Sub = ctk.CTkEntry(root1,placeholder_text="Enter Subject-Code", 
                                        width=360,
                                        height=40,
                                        font=ctk.CTkFont(size=15,weight="bold"),
                                        border_color="#3468c9",
                                        corner_radius=15)#check box
        code_Sub.place(x=520,y=292)

        lable6 = ctk.CTkLabel(root1,text="Enter Phone Number",
                                        font=ctk.CTkFont(size=20,weight="bold"))
        lable6.place(x=110,y=355) 
        phone = ctk.CTkEntry(root1,placeholder_text="Enter Phone Number", 
                                        width=360,
                                        height=40,
                                        font=ctk.CTkFont(size=15,weight="bold"),
                                        border_color="#3468c9",
                                        corner_radius=15)#check box
        phone.place(x=110,y=392)

        lable7 = ctk.CTkLabel(root1,text="Enter Password",
                                        font=ctk.CTkFont(size=20,weight="bold"))
        lable7.place(x=520,y=355) 
        passw = ctk.CTkEntry(root1,placeholder_text="Enter Password", 
                                        width=360,
                                        height=40,
                                        font=ctk.CTkFont(size=15,weight="bold"),
                                        border_color="#3468c9",
                                        corner_radius=15)#check box
        passw.place(x=520,y=392)

        btnx = ctk.CTkButton(root1,text="Register",
                                    height=35,
                                    font=ctk.CTkFont(size=15,weight="bold"),
                                    command=register
                                    )
        btnx.place(y=492,x=420)
        root.mainloop()







    def register_page_stu():

            def register():
                username = usn.get()
                password = passw.get()
                dept = optionmenu_1.get()
                name = namee.get()
                phnumb =phone.get()
                roll = code_Sub.get()
                if len(username) == 0 or len(password) == 0 or len(name) == 0 or len(dept) == 0 or len(roll) == 0 or len(phnumb) == 0:
                    messagebox.showerror(title="Sign Up",message="Plese Fill All Details")
                else:
                    try:
                        user = auth.create_user_with_email_and_password(email=username,password=password)
                        data = {"Role":"Student","Name":name,"Depertment":dept,"Roll Number":roll,"Email":username,"Password":password,"Phone Number":phnumb}
                        # db.child("Students").child(roll).set(data)
                        db.collection("Student").document(roll).set(data)  #for add data in cloud store

                        messagebox.showinfo(title="Sign Up",message="Registration Successful, Now you can login with your email and password !")
                        root.destroy()
                    except:
                        messagebox.showerror(title="Sign Up",message="Invalid email or password, Try Again !")
                        root.destroy()

                        
            root = ctk.CTk()
            root.geometry("1100x700")
            root.title("Register Student")
            root1 = ctk.CTkFrame(root,height=600,width=1000)
            root1.pack(pady=50)
            label1 = ctk.CTkLabel(root1,text="Register Student",font=ctk.CTkFont(size=40,weight="bold"))
            label1.place(x=330,y=45)

            def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


            mark = ctk.CTkLabel(root1, 
                        font = ('calibri', 20, 'bold'))
            mark.place(x=870,y=65)
            time()



            lable2 = ctk.CTkLabel(root1,text="Name",
                                            font=ctk.CTkFont(size=20,weight="bold"))
            lable2.place(x=115,y=155) 
            namee = ctk.CTkEntry(root1,placeholder_text="Enter Name", 
                                            width=360,
                                            height=40,
                                            font=ctk.CTkFont(size=15,weight="bold"),
                                            border_color="#3468c9",
                                            corner_radius=15)#check box
            namee.place(x=110,y=192)


            lable3 = ctk.CTkLabel(root1,text="Enter Email",
                                            font=ctk.CTkFont(size=20,weight="bold"))
            lable3.place(x=520,y=155) 

            usn = ctk.CTkEntry(root1,placeholder_text="Enter Email", 
                                            width=360,
                                            height=40,
                                            font=ctk.CTkFont(size=15,weight="bold"),
                                            border_color="#3468c9",
                                            corner_radius=15)
            usn.place(x=520,y=192)



            lable4 = ctk.CTkLabel(root1,text="Enter Department",
                                            font=ctk.CTkFont(size=20,weight="bold"))
            lable4.place(x=115,y=255) 

            optionmenu_1 = ctk.CTkOptionMenu(root1,dynamic_resizing=False,
                                                            values=["CSE","IT","ECE","EE","ME","CE"],
                                                            width=360,
                                                            height=35,
                                                            dropdown_hover_color="#2f4978",
                                                            font=ctk.CTkFont(size=18,weight="bold"),
                                                            dropdown_font=ctk.CTkFont(size=20,weight="bold"))
            optionmenu_1.place(x=110,y=292)

            lable5 = ctk.CTkLabel(root1,text="Enter Roll-Number",
                                            font=ctk.CTkFont(size=20,weight="bold"))
            lable5.place(x=520,y=255) 
            code_Sub = ctk.CTkEntry(root1,placeholder_text="Enter Roll-Number", 
                                            width=360,
                                            height=40,
                                            font=ctk.CTkFont(size=15,weight="bold"),
                                            border_color="#3468c9",
                                            corner_radius=15)#check box
            code_Sub.place(x=520,y=292)

            lable6 = ctk.CTkLabel(root1,text="Enter Phone Number",
                                            font=ctk.CTkFont(size=20,weight="bold"))
            lable6.place(x=110,y=355) 
            phone = ctk.CTkEntry(root1,placeholder_text="Enter Phone Number", 
                                            width=360,
                                            height=40,
                                            font=ctk.CTkFont(size=15,weight="bold"),
                                            border_color="#3468c9",
                                            corner_radius=15)#check box
            phone.place(x=110,y=392)

            lable7 = ctk.CTkLabel(root1,text="Enter Password",
                                            font=ctk.CTkFont(size=20,weight="bold"))
            lable7.place(x=520,y=355) 
            passw = ctk.CTkEntry(root1,placeholder_text="Enter Password", 
                                            width=360,
                                            height=40,
                                            font=ctk.CTkFont(size=15,weight="bold"),
                                            border_color="#3468c9",
                                            corner_radius=15)#check box
            passw.place(x=520,y=392)

            btnx = ctk.CTkButton(root1,text="Register",
                                        height=35,
                                        font=ctk.CTkFont(size=15,weight="bold"),
                                        command=register
                                    )
            btnx.place(y=492,x=420)
            root.mainloop()

    # -----------------------------------------------------------------------------------------------------Add Profile Page
    def add_page(event):
            delete_page()
            def register_choice():
                if(optionmenu_1.get()=="Student"):
                     register_page_stu()
                else:
                     register_page_fac()
                           
            frame1 = ctk.CTkFrame(main_frame,width=1100,height=750)
            img1 = PhotoImage(file="logo1 (1).png")
            label = ctk.CTkLabel(frame1, image=img1,text="",height=10,width=10)
            label.place(x=35,y=16)



            def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


            mark = ctk.CTkLabel(frame1, 
                        font = ('calibri', 20, 'bold'))
            mark.place(x=910,y=26)
            time()


            lablex= ctk.CTkLabel(frame1,text="Welcome",
                                font=ctk.CTkFont(weight="bold",size=40,family="Verdana"))
            lablex.place(x=100,y=15)

            labley= ctk.CTkLabel(frame1,text=user_data["Name"],
                                font=ctk.CTkFont(weight="bold",size=27,family="Helvetica"))
            labley.place(x=30,y=70)




            l1 =ctk.CTkLabel(frame1,text="Add Profile",
                            font=ctk.CTkFont(weight="bold",size=35,family="Verdana"))
            l1.place(x=425,y=160)



            l1 =ctk.CTkLabel(frame1,text="Add Profile Manually",
                            font=ctk.CTkFont(weight="bold",size=18,family="Verdana"))
            l1.place(x=90,y=240)


            l2 =ctk.CTkLabel(frame1,text="Add :  ",
                            font=ctk.CTkFont(weight="bold",size=20))
            l2.place(x=80,y=290)

            optionmenu_1 = ctk.CTkOptionMenu(frame1,dynamic_resizing=False,
                                                values=["Faculty", "Student"],
                                                font=ctk.CTkFont(size=15,weight="bold"),
                                                dropdown_font=ctk.CTkFont(size=15,weight="bold"))
            optionmenu_1.place(x=150,y=292)


            bt2 =ctk.CTkButton(frame1,
                            text="Launch",
                            font=ctk.CTkFont(size=13,weight="bold"),
                            width=70,
                            command=register_choice)
            bt2.place(x=200,y=340)


            l3 =ctk.CTkLabel(frame1,text="Import Excel Sheet",
                            font=ctk.CTkFont(weight="bold",size=18,family="Verdana"))
            l3.place(x=720,y=240)

            l4 =ctk.CTkLabel(frame1,text="Import Facultys :",
                            font=ctk.CTkFont(weight="bold",size=20))
            l4.place(x=680,y=290)


            bt3 =ctk.CTkButton(frame1,
                            text="Import",
                            font=ctk.CTkFont(size=13,weight="bold"),
                            width=70,command=select_excel_file_fac)
            bt3.place(x=880,y=290)


            l4 =ctk.CTkLabel(frame1,text="Import Students :",
                            font=ctk.CTkFont(weight="bold",size=20))
            l4.place(x=680,y=340)


            bt3 =ctk.CTkButton(frame1,
                            text="Import",
                            font=ctk.CTkFont(size=13,weight="bold"),
                            width=70,command=select_excel_file_stu)
            bt3.place(x=880,y=340)

            frame1.pack(pady=20)


    
    def sched(event):
        delete_page()
        root = ctk.CTkFrame(main_frame,width=1100,height=750)

        ctk.set_appearance_mode("Light")
        root1 = ctk.CTkFrame(root,height=40,width=130)

        root1.place(x=10,y=10)

        root2 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root2,text="Monday",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=25,y=5)
        root2.place(x=160,y=10)

        root3 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root3,text="Tuesday",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=20,y=5)
        root3.place(x=310,y=10)

        root4 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root4,text="Wednesday",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=10,y=5)
        root4.place(x=460,y=10)

        root5 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root5,text="Thursday",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=20,y=5)
        root5.place(x=610,y=10)

        root6 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root6,text="Friday",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=35,y=5)
        root6.place(x=760,y=10)
        # =================================================================================

        root7 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root7,text="IT Workshop",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root7,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root7.place(x=760,y=60)
        root7.configure(fg_color="#33DF9B")


        root8 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root8,text="Engneering Biology",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root8,text="(BSC-BS-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root8.place(x=610,y=60)
        root8.configure(fg_color="#33DF9B")



        root9 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root9,text="Digital Logic",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root9,text="(ECE-CSE-302)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root9.place(x=460,y=60)
        root9.configure(fg_color="#33DF9B")


        root10 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root10,text="Analog Electronic ",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root10,text="(ECE-EC-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root10.place(x=310,y=60)
        root10.configure(fg_color="#33DF9B")


        root11 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root11,text="Data Structure & Algorithms",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=6,y=0)
        label1 = ctk.CTkLabel(root11,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=35,y=17)
        root11.place(x=160,y=60)
        root11.configure(fg_color="#33DF9B")



        root12 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root12,text="10:00-10:50",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root12.place(x=10,y=60)


        # ================================================================================


        root13 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root13,text="Engneering Biology",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root13,text="(BSC-BS-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root13.place(x=760,y=110)
        root13.configure(fg_color="#33DF9B")


        root14 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root14,text="Digital Logic",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root14,text="(ECE-CSE-302)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root14.place(x=610,y=110)
        root14.configure(fg_color="#33DF9B")


        root15 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root15,text="IT Workshop",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root15,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root15.place(x=460,y=110)
        root15.configure(fg_color="#33DF9B")


        root16 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root16,text="Data Structure & Algorithms",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=6,y=0)
        label1 = ctk.CTkLabel(root16,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=35,y=17)
        root16.place(x=310,y=110)
        root16.configure(fg_color="#33DF9B")


        root17 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root17,text="Analog Electronic ",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root17,text="(ECE-EC-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root17.place(x=160,y=110)
        root17.configure(fg_color="#33DF9B")


        root18 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root18,text="10:50-11:40",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root18.place(x=10,y=110)



        # =======================================================================================

        root19 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root19,text="Digital Logic",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root19,text="(ECE-CSE-302)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root19.place(x=760,y=160)
        root19.configure(fg_color="#33DF9B")


        root20 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root20,text="Analog Electronic ",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root20,text="(ECE-EC-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root20.place(x=610,y=160)
        root20.configure(fg_color="#33DF9B")


        root21 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root21,text="Differential Calculus",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root21,text="(BSC-M-302)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root21.place(x=460,y=160)
        root21.configure(fg_color="#EC706C")


        root22 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root22,text="Data Structure & Algorithms",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=6,y=0)
        label1 = ctk.CTkLabel(root22,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=35,y=17)
        root22.place(x=310,y=160)
        root22.configure(fg_color="#33DF9B")


        root23 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root23,text="Engneering Biology",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root23,text="(BSC-BS-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root23.place(x=160,y=160)
        root23.configure(fg_color="#33DF9B")


        root24 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root24,text="11:40-12:30",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root24.place(x=10,y=160)

        # =======================================================================================

        root25 = ctk.CTkFrame(root,height=40,width=130)
        root25.place(x=760,y=210)

        root26 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root26,text="Differential Calculus",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root26,text="(BSC-M-302)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root26.place(x=610,y=210)
        root26.configure(fg_color="#33DF9B")


        root27 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root27,text="Linguistic & Oral Communication",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=6,y=0)
        label1 = ctk.CTkLabel(root27,text="(HSM-HU-381)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=35,y=17)
        root27.place(x=460,y=210)
        root27.configure(fg_color="#33DF9B")


        root28 = ctk.CTkFrame(root,height=40,width=130)
        root28.place(x=310,y=210)

        root29 = ctk.CTkFrame(root,height=40,width=130)
        root29.place(x=160,y=210)

        root30 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root30,text="12:30-13:20",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root30.place(x=10,y=210)

        # =======================================================================================


        root31 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root31,text="K",font=ctk.CTkFont(size=30,weight="bold"))
        label.place(x=55,y=3)
        root31.place(x=760,y=260)
        root31.configure(fg_color="#e0ce00")

        root32 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root32,text="A",font=ctk.CTkFont(size=30,weight="bold"))
        label.place(x=55,y=3)
        root32.place(x=610,y=260)
        root32.configure(fg_color="#e0ce00")


        root33 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root33,text="E",font=ctk.CTkFont(size=30,weight="bold"))
        label.place(x=55,y=3)
        root33.place(x=460,y=260)
        root33.configure(fg_color="#e0ce00")


        root34 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root34,text="R",font=ctk.CTkFont(size=30,weight="bold"))
        label.place(x=55,y=3)
        root34.place(x=310,y=260)
        root34.configure(fg_color="#e0ce00")


        root35 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root35,text="B",font=ctk.CTkFont(size=30,weight="bold"))
        label.place(x=55,y=3)
        # label.configure(fg_color="black")
        root35.place(x=160,y=260)
        root35.configure(fg_color="#e0ce00")


        root36 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root36,text="13:20-14:10",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root36.place(x=10,y=260)

        # ===========================================================================================


        root37 = ctk.CTkFrame(root,height=40,width=130)
        root37.place(x=760,y=310)

        root38 = ctk.CTkFrame(root,height=40,width=130)
        root38.place(x=610,y=310)

        root39 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root39,text="IT Workshop",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root39,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root39.place(x=460,y=310)
        root39.configure(fg_color="#6ED8EA")


        root40 = ctk.CTkFrame(root,height=40,width=130)
        root40.place(x=310,y=310)

        root41 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root41,text="Analog Electronic ",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root41,text="(ECE-EC-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root41.place(x=160,y=310)
        root41.configure(fg_color="#6ED8EA")


        root42 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root42,text="14:10-15:00",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root42.place(x=10,y=310)

        # ====================================================================================================


        root43 = ctk.CTkFrame(root,height=40,width=130)
        root43.place(x=760,y=360)

        root44 = ctk.CTkFrame(root,height=40,width=130)
        root44.place(x=610,y=360)

        root45 = ctk.CTkFrame(root,height=40,width=130)
        root45.place(x=460,y=360)

        root46 = ctk.CTkFrame(root,height=40,width=130)
        root46.place(x=310,y=360)

        root47 = ctk.CTkFrame(root,height=40,width=130)
        root47.place(x=160,y=360)

        root48 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root48,text="15:00-15:50",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root48.place(x=10,y=360)

        # =======================================================================================


        root49 = ctk.CTkFrame(root,height=40,width=130)
        root49.place(x=760,y=410)

        root50 = ctk.CTkFrame(root,height=40,width=130)
        root50.place(x=610,y=410)

        root51 = ctk.CTkFrame(root,height=40,width=130)
        root51.place(x=460,y=410)

        root52 = ctk.CTkFrame(root,height=40,width=130)
        root52.place(x=310,y=410)

        root53 = ctk.CTkFrame(root,height=40,width=130)
        root53.place(x=160,y=410)

        root54 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root54,text="15:50-16:40",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root54.place(x=10,y=410)

        # =======================================================================================

        root55 = ctk.CTkFrame(root,height=40,width=130)
        root55.place(x=760,y=460)

        root56 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root56,text="Digital Logic",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=30,y=0)
        label1 = ctk.CTkLabel(root56,text="(ECE-CSE-302)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=38,y=17)
        root56.place(x=610,y=460)
        root56.configure(fg_color="#6ED8EA")


        root57 = ctk.CTkFrame(root,height=40,width=130)
        root57.place(x=460,y=460)

        root58 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root58,text="Data Structure & Algorithms",font=ctk.CTkFont(size=7,weight="bold"))
        label.place(x=6,y=0)
        label1 = ctk.CTkLabel(root58,text="(PCC-IT-301)",font=ctk.CTkFont(size=7,weight="bold"))
        label1.place(x=35,y=17)
        root58.place(x=310,y=460)
        root58.configure(fg_color="#6ED8EA")


        root59 = ctk.CTkFrame(root,height=40,width=130)
        root59.place(x=160,y=460)

        root60 = ctk.CTkFrame(root,height=40,width=130)
        label = ctk.CTkLabel(root60,text="16:40-17:30",font=ctk.CTkFont(size=20,weight="bold"))
        label.place(x=6,y=7)
        root60.place(x=10,y=460)

        root.pack(pady=20)



    # ---------------------------------------------------------------------------------
        
    def fac_pref(event):
        delete_page()

        frame1 = ctk.CTkFrame(main_frame,width=1100,height=750)
        # frame.pack(pady=40)
        lable10 = ctk.CTkLabel(frame1,text="Select Faculty - ",
                                        font=ctk.CTkFont(size=25,weight="bold"))
        lable10.place(x=60,y=102)

        optionmenu_2 = ctk.CTkOptionMenu(frame1,dynamic_resizing=False,
                                                        values=["APR","CK","PDE","IC","CT","PPS","AR","KKC","KGH","PSC","SPJ","NSN","ND","ACH","PKS"],
                                                        width=160,
                                                        height=30,
                                                        dropdown_hover_color="#2f4978",
                                                        font=ctk.CTkFont(size=20,weight="bold"),
                                                        dropdown_font=ctk.CTkFont(size=20,weight="bold"))
        optionmenu_2.place(x=310,y=102)




        lable1 = ctk.CTkLabel(frame1,text="When would you like to take leave?",
                                        font=ctk.CTkFont(size=25,weight="bold"))
        lable1.place(x=140,y=40)


        lable2 = ctk.CTkLabel(frame1,text="Select Time Slot - ",
                                        font=ctk.CTkFont(size=25,weight="bold"))
        lable2.place(x=60,y=170)  

        # whwn woud u like to take leave select working day
        # ----------------------------------------------------------
        var = ctk.StringVar(value="option1")
        rb1 = ctk.CTkRadioButton(frame1, text="1st Half", variable=var, value="option1")
        rb2 = ctk.CTkRadioButton(frame1, text="2nd Half", variable=var, value="option2")
        rb3 = ctk.CTkRadioButton(frame1, text="Full Day", variable=var, value="option3")
        rb1.place(x=310,y=175)
        rb2.place(x=410,y=175)
        rb3.place(x=510,y=175)
        # ------------------------------------------------------------------

        lable3 = ctk.CTkLabel(frame1,text="Select Working day - ",
                                        font=ctk.CTkFont(size=25,weight="bold"))
        lable3.place(x=60,y=230)  




        optionmenu_1 = ctk.CTkOptionMenu(frame1,dynamic_resizing=False,
                                                        values=["Monday","Tuesday","Wednesday","Thursday","Friday"],
                                                        width=160,
                                                        height=30,
                                                        dropdown_hover_color="#2f4978",
                                                        font=ctk.CTkFont(size=20,weight="bold"),
                                                        dropdown_font=ctk.CTkFont(size=20,weight="bold"))
        optionmenu_1.place(x=310,y=232)


        btn = ctk.CTkButton(frame1,text="Submit",
                            font=ctk.CTkFont(size=18,weight="bold")
                            )
        btn.place(x=270,y=310)
        frame1.pack(pady=20)

# --------------------------------------------------------------------------------------------



# ----------------------------------------------------------------------------------------------------
    def gen_schd(event):
            delete_page()
            frame1 = ctk.CTkFrame(main_frame,width=1100,height=750)
            img1 = PhotoImage(file="logo1 (1).png")
            label = ctk.CTkLabel(frame1, image=img1,text="",height=10,width=10)
            label.place(x=35,y=16)



            def time():
                string = strftime('%H:%M:%S %p')
                mark.configure(text = string)
                mark.after(1000, time)


            mark = ctk.CTkLabel(frame1, 
                        font = ('calibri', 20, 'bold'))
            mark.place(x=910,y=16)
            time()


            lablex= ctk.CTkLabel(frame1,text="Welcome",
                                font=ctk.CTkFont(weight="bold",size=40,family="Verdana"))
            lablex.place(x=100,y=15)

            labley= ctk.CTkLabel(frame1,text=user_data["Name"],
                                font=ctk.CTkFont(weight="bold",size=27,family="Helvetica"))
            labley.place(x=30,y=70)

            lablex= ctk.CTkButton(frame1,text="Generate Schedule",
                                font=ctk.CTkFont(weight="bold",size=20))
            lablex.place(x=420,y=175)

            frame1.pack(pady=20)

# ----------------------------------------------------------------------Side Bar
    root = ctk.CTk()
    root.geometry("1300x550")
    # root.minsize(width=900,height = 550)
    # root.maxsize(width=900,height = 550)
    root.title("InteliSched")

    option_frame = ctk.CTkFrame(root,border_color="#101B46")


    home_btn = ctk.CTkLabel(option_frame,text="Admin Details",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250)
    home_btn.place(y=40)
    home_btn.bind("<Button-1>",home_page)


    home_btn1 = ctk.CTkLabel(option_frame,text="Update Profile",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250)
    home_btn1.place(y=100)
    home_btn1.bind("<Button-1>",edite_page)

    home_btn1x = ctk.CTkLabel(option_frame,text="Add Profile",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250)
    home_btn1x.place(y=160)
    home_btn1x.bind("<Button-1>",add_page)

    home_btn1xx = ctk.CTkLabel(option_frame,text="Faculty Preference",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250)
    home_btn1xx.place(y=220)
    home_btn1xx.bind("<Button-1>",fac_pref)


    home_btn2  = ctk.CTkLabel(option_frame,text="View Schedule",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250,
                                    )
    home_btn2.place(y=280)
    home_btn2.bind("<Button-1>",sched)



    home_btn2x  = ctk.CTkLabel(option_frame,text="Generate Schedule",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250)
    home_btn2x.place(y=340)
    home_btn2x.bind("<Button-1>",gen_schd)


    home_btn3 = ctk.CTkLabel(option_frame,text="View Notice",
                                    font=ctk.CTkFont(weight="bold",size=25),
                                    corner_radius=25,width=250)
    home_btn3.place(y=400)
    home_btn3.bind("<Button-1>",notice_page)



    home_btn4 = ctk.CTkButton(option_frame,text="Log Out",
                                    font=ctk.CTkFont(weight="bold",size=20),
                                    corner_radius=5,
                                    width=250,
                                    height=35,
                                    command=log_out)
    home_btn4.place(y=450)


    option_frame.pack(side=ctk.LEFT)
    option_frame.configure(width=250,height = 750)


# ==================================================================================Opening Page
    main_frame = ctk.CTkFrame(root,bg_color="#333740")
    main_frame.pack(side=ctk.LEFT)
    main_frame.pack_propagate(False)
    main_frame.configure(width=1150,height=750)
    img1 = PhotoImage(file="logo1 (1).png")
    label = ctk.CTkLabel(main_frame, image=img1,text="",height=10,width=10)
    label.place(x=35,y=16)



    def time():
        string = strftime('%H:%M:%S %p')
        mark.configure(text = string)
        mark.after(1000, time)


    mark = ctk.CTkLabel(main_frame, 
                            font = ('calibri', 20, 'bold'))
    mark.place(x=910,y=25)
    time()
    lablex1= ctk.CTkLabel(main_frame,text="IntelliSched",
                                    font=ctk.CTkFont(weight="bold",size=40,family="Helvetica"))
    lablex1.place(x=100,y=15)

    lablex2= ctk.CTkLabel(main_frame,text="Welcome",
                                    font=ctk.CTkFont(weight="bold",size=55,family="Verdana"))
    lablex2.place(x=380,y=155)

    lablex2= ctk.CTkLabel(main_frame,text=user_data["Name"],
                                    font=ctk.CTkFont(weight="bold",size=35))
    lablex2.place(x=427,y=230)

    root.mainloop()



if __name__ == "__main__":
    handel_admin_details("1")    